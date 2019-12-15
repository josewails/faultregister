from django.shortcuts import render
from django.views import View
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.template.loader import render_to_string
from django.db.models import Count

from .models import Supplier, FaultSource, SubmissionProgress, Submission, Classification
from core.mixins import GeneralMixins
from core.models import Chart

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def write_row(work_sheet, row_num, row_data, column_nums, bold_row=False):
    for column_num in column_nums:

        cell = work_sheet.cell(row=row_num, column=column_num)

        if bold_row:
            cell.value = row_data[column_num - 1] if row_data[column_num - 1] else 'MISSING'
            cell.font = Font(bold=True)

        else:
            cell.value = row_data[column_num - 1] if row_data[column_num - 1] else 'MISSING'

        if not row_data[column_num - 1]:
            cell.fill = PatternFill(bgColor="B22222", fill_type="lightGray")


all_fields = ['submission_id', 'machine_shop_update', 'quality_update', 'stores_update', 'cs_stores_update',
              'time_to_close', 'originator__first_name', 'originator__last_name',
              'originator__area__name', 'fault_source__name',
              'time_in_quality', 'time_in_stores', 'time_in_cs_stores',
              'scan_qr_code', 'supervisor__name', 'supervisor__email', 'input_start_date', 'part_number', 'quantity',
              'total_part_cost', 'order_number', 'rework_quantity', 'scrap_quantity', 'stock_quantity',
              'to_be_reworked_on_site', 'notification_recipient__name', 'part_description', 'bin_number',
              'unit_part_cost',
              'total_ncr_cost', 'classification__name',
              'production_order_number', 'machine_number', 'machine_type__letter', 'machine_classification__name',
              'test_type__name', 'vendor_id', 'supplier__name', 'supplier__email',
              'issue_details__part_or_assembly_image',
              'issue_details__part_or_assembly_detail', 'issue_details__category__name',
              'issue_details__accurate_description',
              'issue_details__exact_description', 'previous_area__name', 'current_area__name', 'sap_updated',
              'rework_time',
              'serial_number', 'create_an_ncr', 'caq_ncr', 'status__name', 'defect_origin_area__name', 'minutes_wasted',
              'remark', 'eight_d_requested', 'eight_d_supplied', 'last_updated_by', 'progress__name', 'draft']

partial_fields = [
    'scan_qr_code', 'submission_id', 'input_start_date', 'machine_number', 'machine_type__letter',
    'originator__area__name',
    'originator__first_name', 'originator__last_name', 'fault_source__name', 'defect_origin_area__name', 'part_number',
    'rework_quantity', 'total_part_cost', 'part_description', 'classification__name', 'create_an_ncr',
    'machine_classification__name', 'issue_details__part_or_assembly_image',
    'issue_details__part_or_assembly_detail', 'issue_details__category__name', 'issue_details__accurate_description',
    'serial_number', 'minutes_wasted', 'status', 'current_area__name', 'vendor_id', 'production_order_number',
    'progress__name', 'sap_updated', 'total_ncr_cost'
]


def make_title(field):
    return ' '.join([word.strip('_').capitalize() for word in field.split('_')])


class HomeView(View):

    @staticmethod
    def get(request):
        context = dict()
        context['charts'] = Chart.objects.order_by('number')
        context['suppliers'] = Supplier.objects.annotate(
            count=Count('submissions__id')
        ).values(
            'count', 'id', 'name'
        ).order_by('-count')
        context['progress'] = SubmissionProgress.objects.values('id', 'name')
        context['classifications'] = Classification.objects.values('id', 'name')

        liabilities = Submission.objects.filter(
            defect_origin_area__isnull=False
        ).annotate(
            count=Count('defect_origin_area__id')
        ).values(
            'count', 'defect_origin_area__id', 'defect_origin_area__name'
        ).order_by(
            '-count'
        ).distinct()

        context['buhler'] = list()
        for liability in liabilities:
            if liability['defect_origin_area__name'] == 'Supplier':
                context['supplier'] = dict(id=liability['defect_origin_area__id'],
                                           name=liability['defect_origin_area__name'])
            else:
                context['buhler'].append(dict(id=liability['defect_origin_area__id'],
                                              name=liability['defect_origin_area__name']))

        return render(request, 'submissions/home.html', context)


class ExcelDownload(GeneralMixins, View):

    def get(self, request):
        submission_ids = self.get_submission_ids()

        show_all_columns = request.GET.get('showAllColumns', '0')

        if show_all_columns == '1':
            show_all_columns = 1
        else:
            show_all_columns = 0

        fields = all_fields if show_all_columns else partial_fields

        submissions = Submission.objects.select_related(
            'originator'
        ).prefetch_related(
            'originator__area'
        ).select_related(
            'fault_source'
        ).select_related(
            'supervisor'
        ).select_related(
            'supplier'
        ).select_related(
            'previous_area'
        ).select_related(
            'current_area'
        ).select_related(
            'defect_origin_area'
        ).select_related(
            'notification_recipient'
        ).select_related(
            'issue_details'
        ).select_related(
            'classification'
        ).select_related(
            'status'
        ).select_related(
            'progress'
        ).filter(
            id__in=submission_ids
        ).order_by('submission_id').values(*fields)

        current_row = 1
        work_book = Workbook()
        work_sheet = work_book.active
        work_sheet.title = f'Filtered Submissions'
        buffer = BytesIO()

        columns = [_ for _ in range(1, len(fields) + 1)]
        titles = [make_title(field) for field in fields]

        write_row(work_sheet, current_row, titles, columns, bold_row=True)

        current_row += 1
        submissions_data = list()

        for submission in submissions:
            temp_data = list()
            for field in fields:
                temp_data.append(submission[field])
            submissions_data.append(temp_data)

        for data in submissions_data:
            print(f"Writing {current_row}")
            write_row(work_sheet, current_row, data, columns, bold_row=False)
            current_row += 1

        work_book.save(buffer)

        file_name = f'filtered_submissions_{timezone.now().timestamp()}.xlsx'
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = f'attachment; filename={file_name}'
        response.write(buffer.getvalue())

        return response


class ExcelTable(GeneralMixins, View):

    def get(self, request):

        submission_ids = self.get_submission_ids()
        page_number = request.GET.get('pageNumber', None)
        page_number = int(page_number) if page_number else 1
        show_all_columns = request.GET.get('showAllColumns', '0')

        if show_all_columns == '1':
            show_all_columns = 1
        else:
            show_all_columns = 0

        fields = all_fields if show_all_columns else partial_fields

        submissions = Submission.objects.select_related(
            'originator'
        ).prefetch_related(
            'originator__area'
        ).select_related(
            'fault_source'
        ).select_related(
            'supervisor'
        ).select_related(
            'supplier'
        ).select_related(
            'previous_area'
        ).select_related(
            'current_area'
        ).select_related(
            'defect_origin_area'
        ).select_related(
            'notification_recipient'
        ).select_related(
            'issue_details'
        ).select_related(
            'classification'
        ).select_related(
            'status'
        ).select_related(
            'progress'
        ).filter(
            id__in=submission_ids
        ).order_by('submission_id').values(*fields)

        paginator = Paginator(submissions, 25)
        page_number = min(page_number, paginator.num_pages)
        current_submissions = paginator.page(page_number)

        if page_number < 10:
            pages = [_ for _ in range(1, min(11, paginator.num_pages))]

        elif page_number + 10 > paginator.num_pages:
            pages = [_ for _ in range(page_number, paginator.num_pages + 1)]

        else:
            pages = [_ for _ in range(page_number - 5, page_number + 5)]

        submissions_table_data = list()

        for submission in current_submissions:
            temp_data = list()
            for field in fields:
                temp_data.append(submission[field])

            submissions_table_data.append(temp_data)

        context_one = dict(
            submissions_table_data=submissions_table_data,
            titles=[make_title(field) for field in fields]
        )
        context_two = dict(
            pages=pages,
            has_next=current_submissions.has_next(),
            has_previous=current_submissions.has_previous(),
            next_page=page_number + 1,
            previous_page=page_number - 1,
            page_number=page_number
        )

        return JsonResponse({
            'excel_table': render_to_string('submissions/excel_table.html', context_one),
            'pagination': render_to_string('submissions/pagination.html', context_two)
        })


class SearchSubmissions(View):
    model = Submission
    template_name = 'submissions/part_number_filtered_submissions.html'

    def get(self, request):

        keyword = request.GET.get('keyword', None)
        context = dict()

        if keyword:

            submissions = Submission.objects.order_by('-input_start_date')

            # Give scan QR code the first priority
            scan_qr_code_submissions = submissions.filter(scan_qr_code__icontains=keyword)
            part_number_submissions = submissions.filter(part_number__icontains=keyword)

            if scan_qr_code_submissions.exists():
                context['submissions'] = scan_qr_code_submissions[:10]
                context['result_type'] = 'scan_qr_code'

            # Give part number the second priority

            elif part_number_submissions.exists():
                context['submissions'] = part_number_submissions[:10]
                context['result_type'] = 'part_number'

        return JsonResponse(
            {'html_content': render_to_string('submissions/part_number_filtered_submissions.html', context=context)})
